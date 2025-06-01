import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.reports.models.bulletin_model import Bulletin
from app.academic.models import Enrollment

class ExcelBulletinService:
    def generate_excel_content(self, bulletin: Bulletin):
        wb = Workbook()
        ws = wb.active
        ws.title = "Boletín de Calificaciones"
        
        student = bulletin.student
        trimester = bulletin.trimester
        course_name = "N/A"
        try:
            enrollment = Enrollment.objects.filter(
                student=student, 
                period=trimester.period,
                status='active'
            ).select_related('course').first()
            if enrollment and enrollment.course:
                course_name = enrollment.course.name
        except Enrollment.DoesNotExist:
            pass

        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws.merge_cells('A1:C1')
        ws['A1'] = "Boletín de Calificaciones"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_alignment

        ws.merge_cells('A2:C2')
        ws['A2'] = "Institución Educativa FICCT School"
        ws['A2'].alignment = center_alignment
        
        ws.merge_cells('A3:C3')
        ws['A3'] = f"{trimester.period.name} - {trimester.name}"
        ws['A3'].alignment = center_alignment
        
        ws['A5'] = "Estudiante:"
        ws['A5'].font = bold_font
        ws['B5'] = student.user.get_full_name()
        
        ws['A6'] = "ID Estudiante:"
        ws['A6'].font = bold_font
        ws['B6'] = student.student_id
        
        ws['A7'] = "Curso:"
        ws['A7'].font = bold_font
        ws['B7'] = course_name
        
        ws['A9'] = "Materia"
        ws['A9'].font = bold_font
        ws['A9'].fill = header_fill
        ws['B9'] = "Promedio"
        ws['B9'].font = bold_font
        ws['B9'].alignment = center_alignment
        ws['B9'].fill = header_fill
        
        current_row = 10
        
        for subject_data in bulletin.grades_data.get('subjects', []):
            ws.cell(row=current_row, column=1, value=subject_data.get('subject_name', ''))
            
            average_cell = ws.cell(row=current_row, column=2, value=round(float(subject_data.get('subject_average', 0)), 2))
            average_cell.alignment = center_alignment
            
            current_row += 1
        
        ws.cell(row=current_row + 1, column=1, value="Promedio General").font = bold_font
        overall_cell = ws.cell(row=current_row + 1, column=2, value=round(float(bulletin.overall_average or 0), 2))
        overall_cell.font = bold_font
        overall_cell.alignment = center_alignment
        
        ws.merge_cells(f'A{current_row + 4}:C{current_row + 4}')
        ws[f'A{current_row + 4}'] = f"Generado el: {bulletin.generated_at.strftime('%d/%m/%Y %H:%M') if bulletin.generated_at else 'N/A'}"
        ws[f'A{current_row + 4}'].alignment = center_alignment
        
        ws.merge_cells(f'A{current_row + 5}:C{current_row + 5}')
        ws[f'A{current_row + 5}'] = f"© {trimester.period.start_date.year if trimester.period and trimester.period.start_date else '2025'} FICCT School. Todos los derechos reservados."
        ws[f'A{current_row + 5}'].alignment = center_alignment
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        
        for row in range(9, current_row + 2):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"bulletin_{bulletin.pk}.xlsx"
        return excel_buffer.getvalue(), filename

excel_bulletin_service = ExcelBulletinService()